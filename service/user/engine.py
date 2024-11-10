import os
import json
import time
import importlib
import requests
import shutil
from datetime import datetime, timedelta
import traceback
from collections import namedtuple
import tempfile
import logging

import gspread
from google.oauth2.service_account import Credentials
import xlrd
import openpyxl
from pywinauto.application import Application
from pywinauto import Desktop

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from webdriver_manager.chrome import ChromeDriverManager

import models


MIN_ROW = 2
MAX_COL = 10
COLUMNS = namedtuple("COLUMNS", ["zero", "one", "two", "three", "four", "five", "six", "seven", "eight", "nine"])
CHAT_BOX_URL = "http://127.0.0.1:5001/get-chat-box"
# Paths
target_dir = r'C:\tmp\chrome-data\watsapp-session'

class SBMGoogleSheet:
    # Path to the service account key file
    #SERVICE_ACCOUNT_FILE = r'C:\Users\SHWETHA\OneDrive\Documents\Jaga\Learning\Python\GasOffice\sbm-bcast\service\user\config\g.json'
    # Define the scope
    SCOPES = ["https://www.googleapis.com/auth/spreadsheets.readonly"]
    
    def __init__(self, sheet_url, is_gsheet=False):
        if is_gsheet:
            temp_file_path = ""
            with open(os.path.join("config", "g.json"), 'w') as g_json:
                g_json.write(self.service_passkey)
                temp_file_path = os.path.join("config", "g.json")
            # Authenticate and create a client
            self.creds = Credentials.from_service_account_file(
                temp_file_path,
                scopes=SBMGoogleSheet.SCOPES
            )
            self.client = gspread.authorize(self.creds)
        self.sheet_url = sheet_url
    
    @property
    def service_passkey(self):
        with models.get_msg_db() as conn:
            settings = models.get_settings(
                conn=conn
            )
            return settings[1]

    @property
    def sheet(self):
        with models.get_msg_db() as conn:
            settings = models.get_settings(
                conn=conn
            )
            # Open the Google Sheet using its URL or name
            return self.client.open_by_url(
                self.sheet_url
            ).worksheet(settings[3])

    def fetch_all_records(self):
        # Fetch all records from the sheet
        return self.sheet.get_all_values()

    def fetch_all_records_local(self):
        file_extension = self.sheet_url.split('.')[-1].lower()
        record = []
        print("file_extension ", file_extension)
        if file_extension == 'xlsx':
            workbook = openpyxl.load_workbook(self.sheet_url)
            sheet = workbook.active
            print("sheet ", sheet)
            for row in sheet.iter_rows(min_row=MIN_ROW, max_row=sheet.max_row, max_col=MAX_COL, values_only=True):
                if any(cell is not None and cell != '' for cell in row):
                    print(row)
                    columns = COLUMNS(*row[:10])
                    record.append(columns)
            return record
    
        elif file_extension == 'xls':
            workbook = xlrd.open_workbook(self.sheet_url)
            sheet = workbook.sheet_by_index(0)
            for row_idx in range(1, sheet.nrows):
                col_data = []
                for col_idx in range(sheet.ncols):
                    # Get the value of each cell in the row
                    col_data.append(sheet.cell_value(row_idx, col_idx))
                columns = COLUMNS(*col_data[:10])
                record.append(columns)
            return record
        else:
            raise ValueError("Unsupported file format")

class SBMWatsApp:
    BASE_URL = "https://web.whatsapp.com/"
    CHAT_URL = "https://web.whatsapp.com/send?phone={phone}&text&type=phone_number&app_absent=1"

    def __init__(self):
        self.browser = ''
        self.load_fresh_browser = False #TODO: Change it to True in production

    def focus_our_browser(self):
        windows = Desktop(backend="uia").windows()
        chrome_window_pid = ''
        time.sleep(30)
        for win in windows:
            print(f"Title | {win.window_text()}")
            if "Bcast Bot" in win.window_text():  # Match the title set by Selenium
                chrome_window_pid = win.process_id()
                print(f"Found Chrome window with PID: {chrome_window_pid}")
        
        # Connect to the Chrome window using the PID
        app = Application(backend="uia").connect(process=chrome_window_pid)
        chrome_window = app.top_window()
        # Bring the window to the foreground
        chrome_window.set_focus()
    
    @property
    def browser_timeout(self):
        with models.get_msg_db() as conn:
            settings = models.get_settings(
                conn=conn
            )
            # Open the Google Sheet using its URL or name
            return settings[4]

    def open_browser(self, headless=True):
        chrome_options = Options()
        if headless:
            #os.system("taskkill /IM chrome.exe /F")
            chrome_options.add_argument("--headless")
        chrome_options.add_argument("--start-maximized")
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument(f"user-data-dir={target_dir}")
        browser = webdriver.Chrome(ChromeDriverManager().install(), options=chrome_options)
        
        if not headless:
            browser.execute_script("document.title = 'Bcast Bot';")
            self.focus_our_browser()
        print("Browser ", browser)
        return browser

    def get_chat_box(self):
        response = requests.get(CHAT_BOX_URL)
        print("response ", response,  response.status_code)
        if response.status_code == 200:
            response = response.json()
            return response["chat_box"]
        else:
            print("Failed to get chat box")
            return None
    
    def get_chat_box_settings(self):
        with models.get_msg_db() as conn:
            settings = models.get_settings(
                conn=conn
            )
            print("settings ", settings[2])
            return settings[2]

    def send_message(self, conn, **data):
        print("received data ", data)
        msg_status = "success"
        try:
            self.browser = self.open_browser()
            sheet = SBMGoogleSheet(data.get("google_sheet_url"), is_gsheet=True)
            message = data.get("message")
            for customer in sheet.fetch_all_records()[1:]:
                to_be_filled = MAX_COL - len(customer)
                customer = COLUMNS(*customer) if to_be_filled == 0 else COLUMNS(*customer, *[None] * to_be_filled)
                print("customer ", customer)
                wphone = str(customer.zero)
                name = customer.one
                amount = customer.two
                format_msg = message.format(name, amount)
                try:
                    self.browser.get(SBMWatsApp.CHAT_URL.format(phone=wphone))
                    time.sleep(3)
                    if self.load_fresh_browser:
                        time.sleep(180)  # Wait for the chat to load
                        self.load_fresh_browser = False
                    # Send the text message
                    inp_xpath = self.get_chat_box_settings()
                    
                    print("before input_box ")
                    input_box = WebDriverWait(self.browser, self.browser_timeout).until(EC.presence_of_element_located((By.XPATH, inp_xpath)))

                    print("input_box ", input_box)
                    # Split the message by `\n` and send each part with a line break
                    for line in format_msg.split('\n'):
                        input_box.send_keys(line)
                        input_box.send_keys(Keys.SHIFT, Keys.ENTER)  # Simulate line break
                    input_box.send_keys(Keys.ENTER)
                    print("Message Sent")
                    time.sleep(10)
                    models.update_message_status(conn, wphone, format_msg, at=datetime.now())

                except Exception as _e:
                    models.update_message_status(conn, wphone, format_msg, "failed", at=datetime.now())
                    print(f"Failed to send the message for the {wphone} due to an error {_e}")
                    msg_status = "failed"
                    traceback.print_exc()
                    continue
        except Exception as __e:
            print(f"Failed to send the message due to an error {__e}")
            msg_status = "failed"
            traceback.print_exc()
            self.browser.quit()
            del self.browser
        finally:
            if self.browser:
                self.browser.quit()
                del self.browser
            return msg_status

    def send_message_non_bulk(self, conn, **data):
        msg_status = "success"
        try:
            self.browser = self.open_browser()
            print("file_path ", data.get("file_path"))
            sheet = SBMGoogleSheet(data.get("file_path"))
            message = data.get("message")
            for customer in sheet.fetch_all_records_local():
                print("customer ", customer)
                format_msg = message.format(customer.one, customer.two)
                print("format_msg ", format_msg)
                wphone = ""
                try:
                    wphone = str(customer.zero)
                    models.update_message_status(conn, wphone, format_msg, "progress")
                    self.browser.get(SBMWatsApp.CHAT_URL.format(phone=wphone))
                    time.sleep(3)
                    if self.load_fresh_browser:
                        time.sleep(180)  # Wait for the chat to load
                        self.load_fresh_browser = False
                    # Send the text message
                    inp_xpath = self.get_chat_box_settings()
                    
                    print("before input_box ")
                    input_box = WebDriverWait(self.browser, self.browser_timeout).until(EC.presence_of_element_located((By.XPATH, inp_xpath)))
                    print("input_box ", input_box)
                    # Split the message by `\n` and send each part with a line break
                    for line in format_msg.split('\n'):
                        input_box.send_keys(line)
                        input_box.send_keys(Keys.SHIFT, Keys.ENTER)  # Simulate line break
                    input_box.send_keys(Keys.ENTER)
                    print("Message Sent")
                    time.sleep(10)
                    models.update_message_status(conn, wphone, format_msg, at=datetime.now())
                except Exception as _e:
                    models.update_message_status(conn, wphone, format_msg, "failed", at=datetime.now())
                    print(f"Failed to send the message for the {wphone} due to an error {_e}")
                    msg_status = "failed"
                    continue
        except Exception as __e:
            msg_status = "failed"
            traceback.print_exc()
            self.browser.quit()
            del self.browser
            pass
        finally:
            if self.browser:
                self.browser.quit()
            return msg_status

    def cleanup_destination(self):
        # 1. Remove all files in the target directory
        if os.path.exists(target_dir):
            for root, dirs, files in os.walk(target_dir, topdown=False):
                for name in files:
                    file_path = os.path.join(root, name)
                    try:
                        os.remove(file_path)  # Remove file
                    except Exception as e:
                        print(f"Error removing file {file_path}: {e}")
                for name in dirs:
                    dir_path = os.path.join(root, name)
                    try:
                        os.rmdir(dir_path)  # Remove empty directory
                    except Exception as e:
                        print(f"Error removing directory {dir_path}: {e}")
        else:
            print(f"Target directory {target_dir} does not exist.")
    
    def watsapp_login(self, headless=False):
        self.browser = self.open_browser(headless)
        try:
            print("Browser opened")
            self.browser.get("https://web.whatsapp.com")
            print("Opened watsapp")
            # Wait for manual login or use automation to interact
            inp_xpath = '//*[@id="side"]/div[1]/div/div[2]/div[2]/div/div/p'
            WebDriverWait(self.browser, self.browser_timeout).until(EC.presence_of_element_located((By.XPATH, inp_xpath)))
            # You can continue with your broadcasting script...
        except Exception as e:
            traceback.print_exc()
            self.browser.quit()
            del self.browser
        finally:
            if self.browser:
                self.browser.quit()
                del self.browser


def background_worker():
    while True:
      try:
        #current_app.logger.info("Checking the queued jobs")
        with models.get_msg_db() as conn:
            jobs = models.get_next_job(conn)
        for job in jobs:
          with models.get_msg_db() as conn:
              if job:
                job_id = job[0]
                status = job[1]
                job_name = job[2]
                params = json.loads(job[3]) or {}
                created_at = job[4]
                frequency = int(job[5]) if job[5] else 0   # Assuming frequency is at index 4
                updated_at = datetime.strptime(job[6], "%Y-%m-%d %H:%M:%S.%f")  # Assuming updated_at is at index 5
              else:
                # Sleep for a short time before checking for new jobs
                time.sleep(5)
              # Check if frequency is passed and the job is due for processing
              if (
                  (updated_at <= datetime.now() and status in ["queued"]) or
                  (frequency and updated_at <= datetime.now() - timedelta(days=frequency) and status != "in_progress")
              ):
                  print("Entering... ", job)
                  models.update_job_status(conn, job_id, 'in_progress')
                  package = module = func = None
                  try:
                      # Dynamically import the module and function
                      package_name, module_name, func_name = job_name.rsplit('.')
                      package = importlib.import_module(package_name)
                      module = getattr(package, module_name)
                      func = getattr(module(), func_name)
                      if func(conn, **params) == "failed":
                          models.update_job_status(conn, job_id, 'failed')
                          continue
                      models.update_job_status(conn, job_id, 'success')
                  except Exception as e:
                      models.update_job_status(conn, job_id, 'failed')
                      print(f"Job {job_id} failed with error: {e}")
              else:
                  # Sleep for a short time before checking for new jobs
                  time.sleep(5)
      except Exception as e:
        #current_app.logger.info(e)
        traceback.print_exc()
      time.sleep(5)
